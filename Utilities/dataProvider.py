import openpyxl
import os

script_dir = os.path.dirname(os.path.realpath(__file__))
parent_dir = os.path.abspath(os.path.join(script_dir, os.pardir))
file_path = os.path.join(parent_dir, "ExcelSheet", "testdata.xlsx")

def get_data(sheetName):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheetName]
    totalrows = sheet.max_row
    totalcols = sheet.max_column

    mainList = []

    for i in range(2, totalrows + 1):
        dataList = []
        for j in range(1, totalcols + 1):
            data = sheet.cell(row=i, column=j).value
            dataList.insert(j, data)
        mainList.insert(i, dataList)

    return mainList


def write_data_to_excel(sheetName, *args):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheetName]

    sheet.append(list(args))

    workbook.save(file_path)


def get_last_row_data(sheetName):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheetName]

    # Get the values from the last row
    last_row_data1 = list(sheet.iter_rows(min_row=sheet.max_row, max_row=sheet.max_row, values_only=True))[0]
    return [last_row_data1]
